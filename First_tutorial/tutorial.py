import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

# Load the Excel file into a DataFrame
file_path = 'sales_data.xlsx'  # Replace with your actual file path
df = pd.read_excel(file_path)

# Perform basic data analysis
df['Total Sales'] = df['Quantity'] * df['Price']
total_sales_by_product = df.groupby('Product')['Total Sales'].sum().reset_index()
average_price_by_product = df.groupby('Product')['Price'].mean().reset_index()
top_products = total_sales_by_product.sort_values(by='Total Sales', ascending=False).head(5)

# Create a new Excel workbook and add a worksheet
summary_report_path = 'summary_report.xlsx'  # Replace with your desired file path
wb = Workbook()
ws = wb.active
ws.title = 'Summary Report'

# Write summary statistics to the Excel worksheet
ws['A1'] = 'Product'
ws['B1'] = 'Total Sales'
ws['C1'] = 'Average Price'
for r_idx, row in enumerate(dataframe_to_rows(total_sales_by_product, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
for r_idx, row in enumerate(dataframe_to_rows(average_price_by_product, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 3):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)

# Add a chart for top-performing products
chart = plt.figure(figsize=(10, 6))
sns.barplot(x='Total Sales', y='Product', data=top_products, palette='viridis')
plt.title('Top Performing Products by Total Sales')
plt.xlabel('Total Sales')
plt.ylabel('Product')
plt.tight_layout()

# Save the chart as an image and add it to the Excel worksheet
chart_image_path = 'top_products_chart.png'  # Replace with your desired image path
chart.savefig(chart_image_path)
img = Image(chart_image_path)
ws.add_image(img, 'E2')

# Save the summary report as an Excel file
wb.save(summary_report_path)
wb.close()

# Optionally, you can also convert the Excel file to PDF using external libraries.
# One option is to use the 'pyexcelerate' library for this purpose.
# Example code for converting the Excel file to PDF:
#
# from pyexcelerate import Workbook as pyWorkbook
# wb = pyWorkbook()
# wb.new_sheet("Summary Report", data=summary_stats.values.tolist(), header=summary_stats.columns.tolist())
# wb.save("summary_report.xlsx")
# wb.to_pdf("summary_report.pdf")
